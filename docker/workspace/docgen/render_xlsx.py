#!/usr/bin/env python3
"""render_xlsx.py — spec (stdin JSON) → .xlsx (argv[1]) via openpyxl.

Spec shapes accepted:
  { "sheets": [ { "name", "columns": [str], "rows": [[cell]], "number_formats": {colIdx: "fmt"} } ] }
  { "columns": [str], "rows": [[cell]] }            # single-sheet shorthand
  { "title", "rows" }                                # bare rows

Auto-formatting (deterministic, runs on every sheet — Rule #1b mechanics, not
prose the model has to remember):
  • URLs   → real clickable hyperlinks (https-normalized, blue underline).
  • Emails → mailto: hyperlinks.
  • Phones → consistent US structure, e.g. "(469) 555-0142" / "+1 (…) …".
  • Addresses → a full-address column is SPLIT into separate Street / City /
    State / ZIP columns (USPS-abbreviated, 2-letter state, ZIP kept as text)
    via usaddress, whenever the cells actually parse into components. If a cell
    can't be parsed it lands whole in Street — data is never dropped. A column
    that yields no city/state/zip in any row is left as-is (just cleaned).
Detection is header-driven (a "Phone"/"Website"/"Email"/"Address" column formats
every cell) with a value-pattern fallback for stray URLs/emails/phones in any
column.

openpyxl writes formulas as strings without computed values; the workspace
/docgen endpoint runs `soffice --headless --convert-to xlsx` afterwards when
the spec sets recalc=true so cached values exist (matches Claude's stack).
"""
import json
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

try:  # real US address parser; degrade gracefully if unavailable
    import usaddress
except Exception:  # noqa: BLE001
    usaddress = None

# Calibri everywhere — the boss's pick for every generated document. openpyxl
# already defaults to Calibri 11, but we set it explicitly so the file (and its
# LibreOffice PDF preview) is guaranteed Calibri regardless of the renderer's
# defaults. Carlito (the metric twin) renders it faithfully in the container.
FONT = "Calibri"
LINK_COLOR = "0563C1"  # Excel's hyperlink blue

# ── detection patterns ───────────────────────────────────────────────────────
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[a-z]{2,}$", re.I)
URL_RE = re.compile(r"^(https?://|www\.)\S+$", re.I)
# bare domain like "foo.com" / "foo.co.uk/path" — requires an alphabetic TLD so
# numbers ("3.14") and emails never match.
DOMAIN_RE = re.compile(r"^(?!.*@)[a-z0-9][a-z0-9\-]*(\.[a-z0-9\-]+)*\.[a-z]{2,}(/\S*)?$", re.I)
PHONE_OK_RE = re.compile(r"^[\d\s().+\-]+$")
EXT_RE = re.compile(r"(?:ext|x|extension)\.?\s*(\d+)\s*$", re.I)

STATE_ABBR = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR", "california": "CA",
    "colorado": "CO", "connecticut": "CT", "delaware": "DE", "florida": "FL", "georgia": "GA",
    "hawaii": "HI", "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA",
    "kansas": "KS", "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV", "new hampshire": "NH",
    "new jersey": "NJ", "new mexico": "NM", "new york": "NY", "north carolina": "NC",
    "north dakota": "ND", "ohio": "OH", "oklahoma": "OK", "oregon": "OR", "pennsylvania": "PA",
    "rhode island": "RI", "south carolina": "SC", "south dakota": "SD", "tennessee": "TN",
    "texas": "TX", "utah": "UT", "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY", "district of columbia": "DC",
}
SUFFIX_ABBR = {
    "street": "St", "avenue": "Ave", "boulevard": "Blvd", "drive": "Dr", "road": "Rd",
    "lane": "Ln", "court": "Ct", "place": "Pl", "parkway": "Pkwy", "highway": "Hwy",
    "circle": "Cir", "terrace": "Ter", "trail": "Trl", "square": "Sq", "loop": "Loop",
    "expressway": "Expy", "freeway": "Fwy", "way": "Way",
}
DIR_ABBR = {
    "north": "N", "south": "S", "east": "E", "west": "W", "northeast": "NE",
    "northwest": "NW", "southeast": "SE", "southwest": "SW",
}
OCC_ABBR = {"suite": "Ste", "apartment": "Apt", "unit": "Unit", "building": "Bldg", "floor": "Fl"}


def _title(s):
    """Title-case, preserving short all-caps tokens (directionals, state codes)."""
    out = []
    for w in str(s).split():
        out.append(w if (w.isupper() and len(w) <= 3) else (w[:1].upper() + w[1:].lower()))
    return " ".join(out)


def _abbr(word, table):
    return table.get(str(word).strip().strip(".").lower(), _title(word))


# ── formatters: url / email / phone ──────────────────────────────────────────
def looks_like_email(v):
    return bool(EMAIL_RE.match(str(v).strip()))


def looks_like_url(v):
    s = str(v).strip()
    if looks_like_email(s):
        return False
    return bool(URL_RE.match(s)) or bool(DOMAIN_RE.match(s))


def normalize_url(v):
    s = str(v).strip()
    return s if re.match(r"^https?://", s, re.I) else "https://" + s


def looks_like_phone(v):
    s = str(v).strip()
    if not s:
        return False
    core = EXT_RE.sub("", s).strip()
    digits = re.sub(r"\D", "", core)
    if not (len(digits) == 10 or (len(digits) == 11 and digits[0] == "1")):
        return False
    if not PHONE_OK_RE.match(core):
        return False
    # Require some punctuation so bare integers / IDs aren't mistaken for phones
    # (a real phone column formats those anyway via header detection).
    return bool(re.search(r"[\s().+\-]", core))


def format_phone(v):
    s = str(v).strip()
    ext = ""
    m = EXT_RE.search(s)
    if m:
        ext = m.group(1)
        s = s[:m.start()]
    digits = re.sub(r"\D", "", s)
    if len(digits) == 11 and digits[0] == "1":
        out = f"+1 ({digits[1:4]}) {digits[4:7]}-{digits[7:11]}"
    elif len(digits) == 10:
        out = f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"
    else:
        return str(v).strip()  # not a standard US number — leave untouched
    return f"{out} x{ext}" if ext else out


# ── formatters: address ──────────────────────────────────────────────────────
def _safe_address(s):
    return re.sub(r"\s*,\s*", ", ", " ".join(str(s).split()))


def parse_address_parts(v):
    """Parse a single address into {street, city, state, zip} (strings, may be
    empty). Unparseable input lands whole in `street` so nothing is lost."""
    s = " ".join(str(v).split())
    blank = {"street": "", "city": "", "state": "", "zip": ""}
    if not s:
        return dict(blank)
    if usaddress is None:
        return {**blank, "street": _safe_address(s)}
    try:
        tagged, kind = usaddress.tag(s)
    except Exception:  # noqa: BLE001
        return {**blank, "street": _safe_address(s)}
    if kind not in ("Street Address", "Ambiguous"):
        return {**blank, "street": _safe_address(s)}

    def g(k):
        return str(tagged.get(k, "")).strip().strip(",")

    street = " ".join(p for p in [
        g("AddressNumber"),
        _abbr(g("StreetNamePreDirectional"), DIR_ABBR) if g("StreetNamePreDirectional") else "",
        _title(g("StreetName")),
        _abbr(g("StreetNamePostType"), SUFFIX_ABBR) if g("StreetNamePostType") else "",
        _abbr(g("StreetNamePostDirectional"), DIR_ABBR) if g("StreetNamePostDirectional") else "",
    ] if p)
    if g("OccupancyType") or g("OccupancyIdentifier"):
        occ = " ".join(p for p in [
            _abbr(g("OccupancyType"), OCC_ABBR) if g("OccupancyType") else "",
            g("OccupancyIdentifier"),
        ] if p)
        street = f"{street}, {occ}" if street else occ
    city = _title(g("PlaceName"))
    st = g("StateName")
    st = STATE_ABBR.get(st.lower(), st.upper() if len(st) <= 2 else _title(st))
    zc = g("ZipCode")
    if g("ZipCodePlus4"):
        zc = f"{zc}-{g('ZipCodePlus4')}"
    if not (street or city or st or zc):
        return {**blank, "street": _safe_address(s)}
    return {"street": street, "city": city, "state": st, "zip": zc}


def format_address(v):
    """One clean address string (used when a column isn't split into parts)."""
    p = parse_address_parts(v)
    tail = " ".join(x for x in [p["state"], p["zip"]] if x)
    joined = ", ".join(x for x in [p["street"], p["city"], tail] if x)
    return joined or _safe_address(v)


# ── column typing ────────────────────────────────────────────────────────────
def is_full_address_header(h):
    """A column that holds a whole address (→ split), not a pre-split part."""
    h = str(h or "").strip().lower()
    return any(k in h for k in ("address", "location", "addr"))


def address_prefix(header):
    """Qualifier kept when naming split columns, e.g. 'Mailing Address' → 'Mailing'."""
    base = re.sub(r"(?i)\b(address|location|addr)\b", "", str(header or ""))
    return re.sub(r"\s+", " ", base).strip(" -,")


def col_kind(header):
    h = str(header or "").strip().lower()
    if not h:
        return None
    if any(k in h for k in ("phone", "tel", "mobile", "cell", "fax")):
        return "phone"
    if any(k in h for k in ("website", "url", "homepage", "domain", "web", "site", "link")):
        return "url"
    if "email" in h or "e-mail" in h or h == "mail":
        return "email"
    if any(k in h for k in ("address", "location", "street", "addr")):
        return "address"
    return None


def cell_kind(value):
    """Per-cell fallback for columns without a typed header (no address guess)."""
    if looks_like_email(value):
        return "email"
    if looks_like_url(value):
        return "url"
    if looks_like_phone(value):
        return "phone"
    return None


def apply_format(cell, kind, value, link_font):
    if kind == "url":
        cell.value = str(value).strip()
        cell.hyperlink = normalize_url(value)
        cell.font = link_font
    elif kind == "email":
        cell.value = str(value).strip()
        cell.hyperlink = "mailto:" + str(value).strip()
        cell.font = link_font
    elif kind == "phone":
        cell.value = format_phone(value)
    elif kind == "address":
        cell.value = format_address(value)


def expand_address_columns(cols, rows):
    """Split each full-address column into Street/City/State/ZIP columns when the
    cells parse into real components. Returns (cols, rows, derived_indices)."""
    addr_idxs = [i for i, c in enumerate(cols) if is_full_address_header(c)]
    if not addr_idxs:
        return cols, rows, set()

    parsed = {}
    for i in addr_idxs:
        parts = [parse_address_parts(r[i] if i < len(r) else "") for r in rows]
        # Only split if at least one row actually yields a city/state/zip —
        # otherwise it's likely street-only or mislabeled; leave it alone.
        if any(p["city"] or p["state"] or p["zip"] for p in parts):
            parsed[i] = parts
    if not parsed:
        return cols, rows, set()

    new_cols, derived = [], set()
    for i, c in enumerate(cols):
        if i in parsed:
            pre = address_prefix(c)
            labels = ["Street", "City", "State", "ZIP"]
            for lbl in labels:
                derived.add(len(new_cols))
                new_cols.append(f"{pre} {lbl}".strip())
        else:
            new_cols.append(c)

    new_rows = []
    for ri, r in enumerate(rows):
        nr = []
        for i, c in enumerate(cols):
            if i in parsed:
                p = parsed[i][ri]
                nr.extend([p["street"], p["city"], p["state"], p["zip"]])
            else:
                nr.append(r[i] if i < len(r) else "")
        new_rows.append(nr)
    return new_cols, new_rows, derived


def add_sheet(wb, spec, first):
    name = (spec.get("name") or "Sheet")[:31]
    ws = wb.active if first else wb.create_sheet()
    ws.title = name
    cols = list(spec.get("columns") or [])
    rows = [list(r or []) for r in (spec.get("rows") or [])]
    link_font = Font(name=FONT, size=11, color=LINK_COLOR, underline="single")

    # Split full-address columns into Street/City/State/ZIP up front.
    cols, rows, derived = expand_address_columns(cols, rows)
    col_kinds = [None if i in derived else col_kind(c) for i, c in enumerate(cols)]

    row_cursor = 1
    if cols:
        for c, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c, value=str(col))
            cell.font = Font(name=FONT, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="333333")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.freeze_panes = "A2"
        row_cursor = 2

    for row in rows:
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=row_cursor, column=c, value=val)
            cell.font = Font(name=FONT, size=11)
            # Wrap long text instead of letting one cell balloon a column —
            # keeps both the live grid AND the PDF preview tidy.
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if val not in (None, ""):
                kind = col_kinds[c - 1] if c - 1 < len(col_kinds) else None
                if kind is None:
                    kind = cell_kind(val)
                if kind:
                    apply_format(cell, kind, val, link_font)
        row_cursor += 1

    # Width auto-fit, capped so a long cell wraps within a sane column rather
    # than stretching the sheet off the page (which wrecks the print preview).
    ncols = len(cols) if cols else (max((len(r) for r in rows), default=0))
    for c in range(1, ncols + 1):
        width = 12
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                width = max(width, min(44, len(str(v)) + 2))
        ws.column_dimensions[get_column_letter(c)].width = width

    # Print/preview layout: the inline preview is a LibreOffice xlsx→PDF render,
    # so without this a wide sheet spills columns onto extra pages or shrinks to
    # tofu. Fit all columns to one page width, repeat the header row on every
    # page, and go landscape once it's wide. The .xlsx stays fully editable —
    # these are just print settings, the same ones a clean deliverable carries.
    if ncols:
        ws.page_setup.orientation = "landscape" if ncols > 6 else "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        if cols:
            ws.print_title_rows = "1:1"


def main():
    out = sys.argv[1]
    spec = json.load(sys.stdin)

    sheets = spec.get("sheets")
    if not sheets:
        sheets = [{
            "name": spec.get("title", "Sheet1"),
            "columns": spec.get("columns", []),
            "rows": spec.get("rows", []),
        }]

    wb = Workbook()
    for i, s in enumerate(sheets):
        add_sheet(wb, s, first=(i == 0))
    wb.save(out)


if __name__ == "__main__":
    main()
