#!/usr/bin/env python3
"""xlsx_to_html.py — <xlsx> (argv[1]) → <html> (argv[2]).

Renders an .xlsx into a clean, horizontally-scrollable HTML table for the
Studio document preview. A spreadsheet should preview AS a spreadsheet — a wide
grid you side-scroll — not as a paginated PDF that chops columns onto separate
pages. Reads with openpyxl so it works for both freshly generated sheets and
backfilled legacy ones, and preserves hyperlinks.
"""
import html
import sys

from openpyxl import load_workbook

HEAD = """<!doctype html>
<html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
  :root { color-scheme: light; }
  * { box-sizing: border-box; }
  html, body { margin: 0; height: 100%; }
  body {
    font-family: Calibri, Carlito, "Segoe UI", system-ui, sans-serif;
    font-size: 13px; color: #1A1A1A; background: #fff;
    overflow: auto; -webkit-overflow-scrolling: touch;
  }
  .sheet { padding: 0 0 24px; }
  h2 { font-size: 14px; font-weight: 600; margin: 16px 12px 6px; color: #1F3864; }
  table { border-collapse: collapse; white-space: nowrap; }
  th, td {
    border: 1px solid #D9DEE2; padding: 6px 11px;
    text-align: left; vertical-align: top;
  }
  thead th {
    position: sticky; top: 0; z-index: 1;
    background: #333; color: #fff; font-weight: 600;
  }
  tbody tr:nth-child(even) { background: #F7F8FA; }
  a { color: #0563C1; text-decoration: none; }
  a:hover { text-decoration: underline; }
</style></head><body>"""


def cell_html(cell):
    v = cell.value
    if v is None:
        return ""
    text = html.escape(str(v))
    link = getattr(cell, "hyperlink", None)
    target = getattr(link, "target", None) if link else None
    # Only linkify safe schemes — a scraped cell could carry a javascript: URI.
    if target and str(target).lower().startswith(("http://", "https://", "mailto:")):
        return f'<a href="{html.escape(str(target), quote=True)}" target="_blank" rel="noopener">{text}</a>'
    return text


def render(path, out):
    wb = load_workbook(path, data_only=True)
    parts = [HEAD]
    multi = len(wb.worksheets) > 1
    for ws in wb.worksheets:
        rows = list(ws.iter_rows())
        if not rows:
            continue
        parts.append('<div class="sheet">')
        if multi:
            parts.append(f"<h2>{html.escape(ws.title)}</h2>")
        parts.append("<table>")
        header, *body = rows
        parts.append("<thead><tr>")
        for c in header:
            parts.append(f"<th>{cell_html(c)}</th>")
        parts.append("</tr></thead><tbody>")
        for r in body:
            parts.append("<tr>")
            for c in r:
                parts.append(f"<td>{cell_html(c)}</td>")
            parts.append("</tr>")
        parts.append("</tbody></table></div>")
    parts.append("</body></html>")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(parts))


if __name__ == "__main__":
    render(sys.argv[1], sys.argv[2])
